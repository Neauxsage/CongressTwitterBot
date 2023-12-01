import os
import random
import shutil
from docx import Document
import tweepy
import openai
import textwrap
import logging
import time
import openpyxl

# Set up logging to print to the terminal
logging.basicConfig(level=logging.INFO)

# Set your API keys
consumer_key = ''# add keys here
consumer_secret = ''# add keys here
access_token = ''# add keys here
access_token_secret = ''# add keys here
openai_api_key = ""# add keys here

# Set up Tweepy client
client = tweepy.Client(consumer_key=consumer_key,
                       consumer_secret=consumer_secret,
                       access_token=access_token,
                       access_token_secret=access_token_secret)
logging.info('Tweepy client set up')

# Set up OpenAI client
openai.api_key = openai_api_key
logging.info('OpenAI client set up')

# Directory where the script is running
base_dir = os.path.dirname(os.path.realpath(__file__))

# Get all Word documents in the 'docs' directory
docs_dir = os.path.join(base_dir, 'texts')
doc_files = [f for f in os.listdir(docs_dir) if f.endswith('.docx')]
logging.info(f'Found {len(doc_files)} docx files in {docs_dir}')

# Select a random Word document
selected_doc_file = random.choice(doc_files)
logging.info(f'Selected file: {selected_doc_file}')

# Open the document and read the text
doc = Document(os.path.join(docs_dir, selected_doc_file))
doc_text = ' '.join([paragraph.text for paragraph in doc.paragraphs])
logging.info('Read text from document')

# Define the chunk size
chunk_size = 3900

# Split the doc text into chunks
doc_text_chunks = textwrap.wrap(doc_text, chunk_size)
logging.info(f'Split text into {len(doc_text_chunks)} chunks')

# List to hold summaries
chunk_summaries = []

# Retry mechanism for OpenAI requests
def openai_request(messages, max_retries=3):
    for attempt in range(max_retries):
        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=messages
            )
            return response['choices'][0]['message']['content']
        except openai.error.RateLimitError:
            if attempt < max_retries - 1:  # No delay on the last attempt
                time.sleep(10)  # Adjust this delay as needed
                continue
            else:
                raise
                
def handle_tweet_error():
    # Delete all documents in the 'texts' directory
    for document in os.listdir(docs_dir):
        os.remove(os.path.join(docs_dir, document))
    logging.info('All documents in the "texts" directory have been deleted')

    # Open the 'links' spreadsheet
    links_file = os.path.join(base_dir, 'links.xlsx')
    wb = openpyxl.load_workbook(links_file)
    sheet = wb['used']

    # Get the last non-empty row number
    last_row = sheet.max_row

    # Check if the last row is empty
    if sheet.cell(row=last_row, column=1).value is None:
        last_row -= 1

    # Check if the last row is not the header
    if last_row > 1:
        # Delete the last row
        sheet.delete_rows(last_row)
        logging.info(f'Deleted row {last_row} from the "used" sheet')

    # Save the workbook
    wb.save(links_file)
    logging.info('Saved the "links" workbook')

# Process each chunk
for i, chunk in enumerate(doc_text_chunks, start=1):
    messages = [
        {"role": "system", "content": "You are a helpful assistant."},
        {"role": "user", "content": chunk},
        {"role": "user", "content": "Craft a very short impartial but compelling tweet about the bill we discussed. Start with an engaging hook, followed by the Bill's Number (without dots) and its official name. Summarize the major points in a concise and fair manner. Incorporate a call to action or an emotional hook and give societal context if relevant. Avoid using quotation marks. Include one or two suitable hashtags. DO not use quotation marks anywhere in the tweet. Do not output how many characters there are. Aim for around 270 characters to fully utilize the space without exceeding Twitter's 280 character limit.Remeber, succinctness, balance, engagement and 280 max character limit should be prioritized."}    ]

    summary = openai_request(messages)
    
    # Add the summary to the list of summaries
    chunk_summaries.append(summary)
    logging.info(f'Summarized chunk {i}/{len(doc_text_chunks)}')

# Combine all summaries into one text
combined_summary = ' '.join(chunk_summaries)
logging.info('Combined chunk summaries')

def generate_tweet_summary(text, max_attempts=12):
    for attempt in range(max_attempts):
        messages = [
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": text},
            {"role": "user", "content": "Craft a very short impartial but compelling tweet about the bill we discussed. Start with an engaging hook, followed by the Bill's Number (without dots) and its official name. Summarize the major points in a concise and fair manner. Incorporate a call to action or an emotional hook and give societal context if relevant. Avoid using quotation marks. Include one or two suitable hashtags. DO not use quotation marks anywhere in the tweet. Do not output how many characters there are. Aim for around 270 characters to fully utilize the space without exceeding Twitter's 280 character limit.Remeber, succinctness, balance, engagement and 280 max character limit should be prioritized."}    ]

        summary = openai_request(messages)

        if len(summary) <= 280:
            return summary
        else:
            logging.info(f'Tweet is too long ({len(summary)} characters). Retrying {max_attempts - attempt - 1} more times.')

    handle_tweet_error()
    return "Unable to generate a suitable summary."
    
# Generate a tweet-sized summary of all summaries
tweet = generate_tweet_summary(combined_summary)
tweet = tweet.replace('"', '')

logging.info(f'Generated tweet: {tweet}')

# Check if the tweet is within the character limit for a tweet
if len(tweet) <= 280 and tweet != "Unable to generate a suitable summary.":
    try:
        # Post a new tweet with the summary
        response = client.create_tweet(text=tweet)
        logging.info('Tweet posted')
    except tweepy.errors.TooManyRequests:
        logging.error('Too many tweets for Tweepy')
        handle_tweet_error()

# Move the document to the 'archive' directory
archive_dir = os.path.join(base_dir, 'archive')
source_file = os.path.join(docs_dir, selected_doc_file)
destination_file = os.path.join(archive_dir, selected_doc_file)

try:
    shutil.move(source_file, destination_file)
    logging.info(f'Moved {selected_doc_file} to archive')
except FileNotFoundError:
    logging.error(f'File {selected_doc_file} not found in the "texts" directory')

# Delete all documents in the 'texts' directory
for document in os.listdir(docs_dir):
    os.remove(os.path.join(docs_dir, document))
logging.info('All documents in the "texts" directory have been deleted')

