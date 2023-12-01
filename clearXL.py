import os
from openpyxl import load_workbook

# Get the current directory
current_directory = os.getcwd()

# Construct the path to the workbook
workbook_path = os.path.join(current_directory, 'links.xlsx')

# Load the workbook
workbook = load_workbook(workbook_path)

# Select the sheet named "text"
sheet = workbook['text']

# Clear all cells in the sheet
sheet.delete_rows(1, sheet.max_row)

# Save the workbook
workbook.save(workbook_path)

# Close the workbook
workbook.close()

print("Cells cleared successfully.")
